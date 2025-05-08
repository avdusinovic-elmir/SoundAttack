import os

import numpy as np
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

folder = "gridSearch/logs"
population = [50, 75, 100, 150]
elite = [0.1, 0.2, 0.3, 0.4]
mutation_range = [0.9, 0.7, 0.5, 0.25]
epsilon_range = [0.9, 0.7, 0.5, 0.25]

d50_4={}
d75_4={}
d100_4={}
d150_4={}
results4 = {
    "50": d50_4,
    "75": d75_4,
    "100": d100_4,
    "150": d150_4
}

d50_8={}
d75_8={}
d100_8={}
d150_8={}
results8 = {
    "50": d50_8,
    "75": d75_8,
    "100": d100_8,
    "150": d150_8
}

folder_content = os.listdir(folder)

for file in folder_content:
    file_split = file.split("_")
    epsilon = file_split[5].split(".")[0]+"."+file_split[5].split(".")[1][:-2]
    elite = str(float(file_split[3])/float(file_split[2]))

    key = elite+"_"+file_split[4]+"_"+epsilon
    result = []

    with open(folder+"/"+file) as f:
        lines = f.readlines()
        perceptual_loss = 0
        epochs = 0
        if "WORD" in lines[-3].strip():
            perceptual_loss = float(lines[-1].split(":")[1].strip())
            epochs = int(lines[-3].strip()[-3:])
        result.append(perceptual_loss)
        result.append(epochs)

    if file_split[1][-1]==str(4):
        results4[file_split[2]].update({key: result})
    else:
        results8[file_split[2]].update({key: result})

# Save the values in an excel file for later use if necessary
# Website used to understand the code:https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/
wb = Workbook()
wb.create_sheet(title="sample4")
wb.create_sheet(title="sample8")
ws4 = wb["sample4"]
ws8 = wb["sample8"]
column = 2
for pop in population:
    ws4.cell(row=1, column=column, value=pop)
    ws4.cell(row=2, column=column, value="perceptual_loss")
    ws4.cell(row=1, column=column+1, value=pop+1)
    ws4.cell(row=2, column=column+1, value="epochs")
    ws8.cell(row=1, column=column, value=pop)
    ws8.cell(row=2, column=column, value="perceptual_loss")
    ws8.cell(row=1, column=column+1, value=pop+1)
    ws8.cell(row=2, column=column + 1, value="epochs")
    column += 2

for key in results4.keys():
    row = 3
    print(key)
    for key2 in results4[key]:
        ws4.cell(row=row, column=1, value=key2)
        position = population.index(int(key))
        result_key = results4[key][key2]
        ws4.cell(row=row, column=2+2*position, value=result_key[0])
        ws4.cell(row=row, column=3+2*position, value=result_key[1])
        row+=1

for key in results8.keys():
    row = 3
    print(key)
    for key2 in results8[key]:
        ws8.cell(row=row, column=1, value=key2)
        position = population.index(int(key))
        result_key = results8[key][key2]
        ws8.cell(row=row, column=2+2*position, value=result_key[0])
        ws8.cell(row=row, column=3+2*position, value=result_key[1])
        row+=1

wb.save(filename="results_logs.xlsx")

df4 = pd.read_excel("results_logs.xlsx", sheet_name="sample4")
df8 = pd.read_excel("results_logs.xlsx", sheet_name="sample8")
plt.figure(figsize=(10,10))
c = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
counter = 0
legend_list = []
# plt.xticks(np.arange(0,1000,step=100))
# plt.yticks(np.arange(0,25,step=1))

# Display all the results in scatterplots

sample4_best_results = []
sample8_best_results = []

for pop in population:
    leg = mpatches.Patch(color=c[counter], label=pop)
    legend_list.append(leg)
    x = df4[pop].values[1:]
    y = df4[pop+1].values[1:]
    for i in range(len(x)):
        if x[i] == 0 and y[i] == 0:
            continue

        if y[i] < 200 and x[i] < 7.5:
            sample4_best_results.append(f"{pop}_{i+1}")
        plt.scatter(y[i], x[i], label=str(pop), marker=f"${i+1}$", s=100, color=c[counter])
    counter+=1
plt.legend(handles=legend_list)
plt.grid(linestyle = '--', linewidth = 0.5)
plt.ylabel("Perceptual loss")
plt.xlabel("Epochs")
plt.title("Sample 4 Grid Search")
plt.savefig("log1.png")
plt.show()

plt.figure(figsize=(10,10))
counter = 0
for pop in population:
    x = df8[pop].values[1:]
    y = df8[pop+1].values[1:]
    for i in range(len(x)):
        if x[i]==0 and y[i]==0:
            continue

        if y[i] < 200 and x[i] < 15:
            # print(str(y[i])+"/"+str(x[i]))
            # print(pop)
            # print(i+1)
            sample8_best_results.append(f"{pop}_{i+1}")
        plt.scatter(y[i], x[i], label=str(pop), marker=f"${i+1}$", s=100, color=c[counter])
    counter+=1
plt.legend(handles=legend_list)
plt.grid(linestyle = '--', linewidth = 0.5)
plt.ylabel("Perceptual loss")
plt.xlabel("Epochs")
plt.title("Sample 8 Grid Search")
plt.savefig("log2.png")
plt.show()

# Try to find the closest results

print(sample4_best_results)
print(sample8_best_results)
counter = 0
temp = 0
for i in range(max(len(sample4_best_results), len(sample8_best_results))):
    if len(sample4_best_results)>len(sample8_best_results):
        if sample4_best_results[i] in sample8_best_results:
            pop = int(sample4_best_results[i].split("_")[0])
            if temp != pop and temp != 0:
                counter+=1
            x4 = df4[pop].values[1:]
            x8 = df8[pop].values[1:]
            y4 = df4[pop + 1].values[1:]
            y8 = df8[pop + 1].values[1:]
            index = int(sample4_best_results[i].split("_")[1])-1

            print(counter)
            mean_y = (y4[index]+y8[index])/2
            mean_x = (x4[index]+x8[index])/2
            plt.scatter(mean_y, mean_x, label=str(pop), marker=f"${index+1}$", s=100, color=c[counter])
            temp = pop
    else:
        if sample8_best_results[i] in sample4_best_results:
            pop = int(sample8_best_results[i].split("_")[0])
            if temp != pop and temp != 0:
                counter+=1
            x4 = df4[pop].values[1:]
            x8 = df8[pop].values[1:]
            y4 = df4[pop + 1].values[1:]
            y8 = df8[pop + 1].values[1:]
            index = int(sample8_best_results[i].split("_")[1])-1

            mean_y = (y4[index] + y8[index]) / 2
            mean_x = (x4[index] + x8[index]) / 2
            plt.scatter(mean_y, mean_x, label=str(pop), marker=f"${index + 1}$", s=100, color=c[counter])
            temp = pop

plt.legend(handles=legend_list)
plt.grid(linestyle = '--', linewidth = 0.5)
plt.ylabel("Perceptual loss")
plt.xlabel("Epochs")
plt.title("Average of the closest results")
plt.savefig("log3.png")
plt.show()
